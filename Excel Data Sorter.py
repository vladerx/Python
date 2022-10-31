import tkinter, openpyxl
from openpyxl import workbook, load_workbook

master = tkinter.Tk();

def check_in_begining(num, target):
	i=0
	while (i<len(num)):
		if (num[i] != target[i]): # checks if target has the same chars as num in the begining
			return False
		i += 1
	if (len(num)<len(target)): # checks if target is another batch of num indicated by '-'
		if (target[i] != '-'):
			return False
	return True

def change_sheet_index(keyword):
	if 'open' in str(keyword):
		if 'AAAA' in str(keyword):
			return 4
		if 'BBBB' in str(keyword):
			return 5
	elif '-50'in str(keyword):
		if 'AAAA' in str(keyword):
			return 2
		if 'BBBB' in str(keyword):
			return 3
	else:
		if 'AAAA' in str(keyword):
			return 0
		if 'BBBB' in str(keyword):
			return 1

def search _main(): #get a phrase and search by it
	word = box.get()

	wb = Workbook() # create new excel file with workbook
	
	wsdef = wb.active #default sheet named after the keyword
	wsdef.title = "AAAA"

	wb.create_sheet('BBBB')
	wb.create_sheet('aged AAAA CL')
	wb.create_sheet('aged BBBB CL')
	wb.create_sheet('aged AAAA OP')
	wb.create_sheet('aged BBBB OP')
	wb.create_sheet('oven')
	wb.create_sheet('notes')
	
	filename = word+' search.xlsx'
	
	wb.save(filename)

	xfile = openpyxl.loadworkbook('')
	sheet = xfile.get_sheet_by_name('sht1')
	
	chemfile = openpyxl.loadworkbook('')
	chemsheet = chemfile.get_sheet_by_name('chem')

	ovenfile = openpyxl.loadworkbook('')
	ovenopsheet = ovenfile.get_sheet_by_name('85%')
	ovenclupsheet = ovenfile.get_sheet_by_name('upper')
	ovenclbotsheet = ovenfile.get_sheet_by_name('bottom')

	ovensheet = [ovenopsheet,ovenclupsheet,ovenclbotsheet]
	
	workb = load_workbook(filename)
	tarsheetcom = workb.get_sheet_by_name('AAAA')
	tarsheetlab = workb.get_sheet_by_name('BBBB')
	tarsheetCAcom = workb.get_sheet_by_name('aged AAAA CL')
	tarsheetCAlab = workb.get_sheet_by_name('aged BBBB CL')
	tarsheetOPcom = workb.get_sheet_by_name('aged AAAA OP')
	tarsheetOPlab = workb.get_sheet_by_name('aged BBBB OP')
	tarsheetOV = workb.get_sheet_by_name('oven')
	tarsheetnote = workb.get_sheet_by_name('notes')

	tarsheet = [tarsheetcom, tarsheetlab, tarsheetCAcom, tarsheetCAlab, tarsheetOPcom, tarsheetOPlab]

	for shet in tarsheet:
		shet.cell(row=1, column=1).value = 'chem number'
		shet.cell(row=1, column=2).value = 'chem name'
		shet.cell(row=1, column=3).value = 'chem weight'
		shet.cell(row=1, column=4).value = 'CT'
	
	tarsheetOV.cell(row=1, column=1).value = 'chem number'
	tarsheetOV.cell(row=1, column=2).value = 'chem name'
	tarsheetOV.cell(row=1, column=3).value = 'outdate'
	tarsheetOV.cell(row=1, column=5).value = 'chem number'
	tarsheetOV.cell(row=1, column=6).value = 'chem name'
	tarsheetOV.cell(row=1, column=7).value = 'outdate'

	tarsheetnote.cell(row=1, column=1).value = 'chem types rows' 
	tarsheetnote.cell(row=2, column=1).value = 'chem file rows' 
	tarsheetnote.cell(row=3, column=1).value = 'oven open rows' 
	tarsheetnote.cell(row=4, column=1).value = 'oven upper rows' 
	tarsheetnote.cell(row=5, column=1).value = 'oven bottom rows'

	run = True

	i = 0
	fileindx = [1,1,1,1,1,1] 
	indx = 0
	remnull = nonull.get()
	while (run): #get chem number by keyword
		i += 1
		chemname = sheet.cell(row=i, column=1).value
		if (chemname == '$'):
			tarsheetnote.cell(row=1, column=2).value = i - 1
			run = False
			chem = False
			break
		if str(word).lower() in str(chemname).lower():
			indx = change_sheet_index(chemname)
			fileindx[indx] += 1
			chemnum = (str(sheet.cell(row=i, column=2).value)).replace(" ","")
			if '-' in str(chemnum):
				fileindx[indx] -= 1
				continue
			if (remnull == 0):
				tarsheet[indx].cell(row=fileindx[indx], column=1).value = str(chemnum)
				tarsheet[indx].cell(row=fileindx[indx], column=2).value = str(chemname)
			chem = True
			k = 0
			ctstr = ""
			ctint = 0
			count = 0
			average = 0
			effchemnum = chemnum
			chemchemweight = 0
			while (chem): #get CT by chem number
				k += 1
				chemchemnum = (str(chemsheet.cell(row=k, column=11).value)).replace(" ","")
				if (chemchemnum == '$'):
					tarsheetnote.cell(row=2, column=1).value = k-1
					if (average != 0):
						tarsheet[indx].cell(row=fileindx[indx], column=1).value = str(effchemnum)
						tarsheet[indx].cell(row=fileindx[indx], column=2).value = str(chemname)
						tarsheet[indx].cell(row=fileindx[indx], column=3).value = str(chemchemweight)
						tarsheet[indx].cell(row=fileindx[indx], column=4).value = ctstr + " ("+str(average)+")"
					else:
						fileindx[indx] -= 1
						chem = False
						break
				if (chemchemnum != none): #skip if empty cell
					if str(effchemnum) in str(chemchemnum): # check if current chem number is part of cell's string
						if (effchemnum == chemchemnum): # check if current chem number is the same as cell's number	
							count += 1
							chemchemweight = chemsheet.cell(row=k,column=8).value
							ctstr += str(chemsheet.cell(row=k,column=1).value)+','
							ctint += float(chemsheet.cell(row=k,column=1).value)
							average = round(ctint/count,2)
						elif (check_in_begining(str(effchemnum),str(chemchemnum))) # check if chem number is at the begining ir its a different number
							indx = change_sheet_index(chemname)
							if (average != 0):
								tarsheet[indx].cell(row=fileindx[indx], column=1).value = str(effchemnum)
								tarsheet[indx].cell(row=fileindx[indx], column=2).value = str(chemname)
								tarsheet[indx].cell(row=fileindx[indx], column=3).value = str(chemchemweight)
								tarsheet[indx].cell(row=fileindx[indx], column=4).value = ctstr + " ("+str(average)+")"
							else:
								if (remnull == 1):
									fileindx[indx] -= 1
								chemname = chemsheet.cell(row=k,column=9).value
								effchemnum = chemchemnum
								fileindx[indx] += 1
								ctstr = ""
								ctint = 0
								count = 0
								average = 0
								k -= 1
					elif str(chemnum) in str(chemchemnum):# check if initial chem number is a part of cell's string
						if (check_in_begining(str(chemnum),str(chemchemnum)))# check if chem number is at the begining or its a different number
							indx = change_sheet_index(chemname)
								if (average != 0):
									tarsheet[indx].cell(row=fileindx[indx], column=1).value = str(effchemnum)
									tarsheet[indx].cell(row=fileindx[indx], column=2).value = str(chemname)
									tarsheet[indx].cell(row=fileindx[indx], column=3).value = str(chemchemweight)
									tarsheet[indx].cell(row=fileindx[indx], column=4).value = ctstr + " ("+str(average)+")"
								else:
									if (remnull == 1):
										fileindx[indx] -= 1
									chemname = chemsheet.cell(row=k,column=9).value
									effchemnum = chemchemnum
									fileindx[indx] += 1
									ctstr = ""
									ctint = 0
									count = 0
									average = 0
									k -= 1
					else: #check if has any ct value and write it
						tarsheet[indx].cell(row=fileindx[indx], column=1).value = str(effchemnum)
						tarsheet[indx].cell(row=fileindx[indx], column=2).value = str(chemname)
						tarsheet[indx].cell(row=fileindx[indx], column=3).value = str(chemchemweight)
						tarsheet[indx].cell(row=fileindx[indx], column=4).value = ctstr + " ("+str(average)+")"
	
	ovenrow = [2,2]
	ovencol = [1,5]
	ovenindx = 0
	ind = 0
	for ovensht in ovensheet:
		ind += 1
		i = 2
		oven = True
		while (oven):
			i += 1
			chemname = ovensht.cell(row=i, column=9).value
			if (chemname == '$'):
				tarsheetnote.cell(row=2+ind, column=2).value = i-1
				oven = False
				break
			if str(word).lower() in str(chemname).lower():
				if 'open' in str(chemname):
					ovindx = 1
				elif '-50' in str(chemname):
					ovindx = 0
				tarsheetOV.cell(row=ovenrow[ovenindx], column=ovencol[ovenindx]).value = ovensht.cell(row=i, column=1).value
				tarsheetOV.cell(row=ovenrow[ovenindx], column=ovencol[ovenindx]+1).value = chemname
				if (ovensht.cell(row=i, column=6).value != None)
					tarsheetOV.cell(row=ovenrow[ovenindx], column=ovencol[ovenindx]+2).value = (ovensht.cell(row=i, column=6).value).strftime('%d-%m-%Y')
				else:
					tarsheetOV.cell(row=ovenrow[ovenindx], column=ovencol[ovenindx]+2).value = ''
					ovenrow[ovindx] += 1
workb.save(filename)
master.geometry('250x75+0+0')
master.title('Excel Data Sorter')
T1 = tkinter.Text(master, height=50, width=45)

box = tkinter.Entry(master)
nonull = tkinter.Intvar()
checkbut = tkinter.Checkbutton(master, text='remove nulls, variable=nonull).grid(row=1, column=0, sticky=tkinter.w, pady=4)
box.place(x=65,y=5)

tkinter.Button(master, text='generate', command=search_main).grid(row=0, column=0,sticky=tkinter.w,pady=4)

tkinter.mainloop()
						
										
							
								
							
			
							
								

		
